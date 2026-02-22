#!/usr/bin/env python3
"""从 style_spread.xlsx 读取数据，生成三个 HTML 看板：
1. 中证红利-科创50 轧差净值
2. 微盘股-中证全指 轧差净值
3. 双创等权净值
"""
import json, os, sys

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

BASE = os.path.expanduser("~/Desktop/size_spread")
wb = openpyxl.load_workbook(os.path.join(BASE, "style_spread.xlsx"))

# === 读 Sheet1 风格轧差 ===
ws1 = wb['风格轧差']
rows1 = list(ws1.iter_rows(min_row=2, values_only=True))

# 列索引: 0=日期, 1=红利chg, 2=科创chg, 3=红利-科创spread, 4=红利-科创nav
#          5=微盘chg, 6=全指chg, 7=微盘-全指spread, 8=微盘-全指nav
#          9=2000chg, 10=300chg, 11=2000-300spread, 12=2000-300nav

pairs = [
    {
        'name': '中证红利 - 科创50',
        'filename': 'dividend_vs_star50.html',
        'label1': '中证红利', 'label2': '科创50',
        'spread_col': 3, 'nav_col': 4,
        'color': '#e67e22',
        'desc': '正值=红利跑赢科创，负值=科创跑赢红利',
    },
    {
        'name': '微盘股 - 中证全指',
        'filename': 'micro_vs_allA.html',
        'label1': '微盘股', 'label2': '中证全指',
        'spread_col': 7, 'nav_col': 8,
        'color': '#9b59b6',
        'desc': '正值=微盘跑赢全A，负值=全A跑赢微盘',
    },
]

def fmt_date(d):
    s = str(d)
    return f"{s[4:6]}/{s[6:8]}"

def gen_spread_html(pair, rows):
    dates_raw = [str(r[0]) for r in rows if r[pair['nav_col']] is not None]
    dates = [fmt_date(d) for d in dates_raw]
    navs = [round(float(r[pair['nav_col']]), 6) for r in rows if r[pair['nav_col']] is not None]
    spreads = [round(float(r[pair['spread_col']]), 4) for r in rows if r[pair['spread_col']] is not None]

    final_nav = navs[-1]
    cum_ret = (final_nav - 1) * 100
    max_spread = max(spreads)
    min_spread = min(spreads)
    avg_spread = sum(spreads) / len(spreads)

    nav_cls = 'pos' if final_nav >= 1 else 'neg'
    cum_cls = 'pos' if cum_ret >= 0 else 'neg'

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>{pair['name']} 轧差策略净值</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
body{{font-family:'PingFang SC',sans-serif;max-width:1000px;margin:40px auto;padding:0 20px;background:#fafafa}}
h2{{text-align:center;color:#333}}
p.sub{{text-align:center;color:#888;font-size:14px;margin-top:-10px}}
.stats{{display:flex;justify-content:center;gap:20px;margin:20px 0;font-size:14px;flex-wrap:wrap}}
.stat{{background:#fff;padding:12px 20px;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,0.1);min-width:100px;text-align:center}}
.stat .label{{color:#888;font-size:12px}}
.stat .value{{font-size:20px;font-weight:bold;margin-top:4px}}
.pos{{color:#e74c3c}} .neg{{color:#2ecc71}}
canvas{{margin-top:20px;background:#fff;border-radius:8px;padding:10px;box-shadow:0 1px 3px rgba(0,0,0,0.1)}}
</style>
</head><body>
<h2>{pair['name']} 轧差策略</h2>
<p class="sub">每日轧差 = {pair['label1']}涨跌幅 - {pair['label2']}涨跌幅 | 归1复利净值 | {dates_raw[0]}~{dates_raw[-1]}</p>
<p class="sub">{pair['desc']}</p>

<div class="stats">
  <div class="stat"><div class="label">最终净值</div><div class="value {nav_cls}">{final_nav:.4f}</div></div>
  <div class="stat"><div class="label">累计收益</div><div class="value {cum_cls}">{cum_ret:+.2f}%</div></div>
  <div class="stat"><div class="label">最大轧差</div><div class="value pos">{max_spread:+.2f}%</div></div>
  <div class="stat"><div class="label">最小轧差</div><div class="value neg">{min_spread:+.2f}%</div></div>
  <div class="stat"><div class="label">平均轧差</div><div class="value">{avg_spread:+.2f}%</div></div>
</div>

<canvas id="navChart" height="100"></canvas>
<canvas id="spreadChart" height="80"></canvas>
<script>
const dates = {json.dumps(dates)};
const navs = {json.dumps(navs)};
const spreads = {json.dumps(spreads)};

new Chart(document.getElementById('navChart'), {{
  type: 'line',
  data: {{labels: dates, datasets: [{{
    label: '归1净值', data: navs,
    borderColor: '{pair["color"]}', backgroundColor: '{pair["color"]}18',
    fill: true, tension: 0.3, pointRadius: 1.5, borderWidth: 2
  }}]}},
  options: {{
    plugins: {{title: {{display:true, text:'{pair["name"]} 轧差策略净值（归1）', font:{{size:14}}}}, legend:{{display:false}}}},
    scales: {{x:{{ticks:{{maxTicksLimit:12}}}}, y:{{title:{{display:true, text:'净值'}}}}}}
  }}
}});

new Chart(document.getElementById('spreadChart'), {{
  type: 'bar',
  data: {{labels: dates, datasets: [{{
    label: '每日轧差(%)', data: spreads,
    backgroundColor: spreads.map(v => v >= 0 ? '{pair["color"]}99' : '#3498db99'),
    borderRadius: 2
  }}]}},
  options: {{
    plugins: {{title: {{display:true, text:'每日涨跌幅轧差（{pair["label1"]} - {pair["label2"]}）', font:{{size:14}}}}, legend:{{display:false}}}},
    scales: {{x:{{ticks:{{maxTicksLimit:12}}}}, y:{{title:{{display:true, text:'%'}}}}}}
  }}
}});
</script>
</body></html>"""

# === 读 Sheet2 双创等权 ===
ws2 = wb['双创等权']
rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
# 列: 0=日期, 1=创业板指chg, 2=科创50chg, 3=等权平均chg, 4=归1净值

def gen_sc_html(rows):
    dates_raw = [str(r[0]) for r in rows if r[4] is not None]
    dates = [fmt_date(d) for d in dates_raw]
    navs = [round(float(r[4]), 6) for r in rows if r[4] is not None]
    avg_chgs = [round(float(r[3]), 4) for r in rows if r[3] is not None]
    cyb_chgs = [round(float(r[1]), 4) for r in rows if r[1] is not None]
    kc_chgs = [round(float(r[2]), 4) for r in rows if r[2] is not None]

    final_nav = navs[-1]
    cum_ret = (final_nav - 1) * 100
    max_chg = max(avg_chgs)
    min_chg = min(avg_chgs)

    nav_cls = 'pos' if final_nav >= 1 else 'neg'
    cum_cls = 'pos' if cum_ret >= 0 else 'neg'

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>双创等权指数净值</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
body{{font-family:'PingFang SC',sans-serif;max-width:1000px;margin:40px auto;padding:0 20px;background:#fafafa}}
h2{{text-align:center;color:#333}}
p.sub{{text-align:center;color:#888;font-size:14px;margin-top:-10px}}
.stats{{display:flex;justify-content:center;gap:20px;margin:20px 0;font-size:14px;flex-wrap:wrap}}
.stat{{background:#fff;padding:12px 20px;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,0.1);min-width:100px;text-align:center}}
.stat .label{{color:#888;font-size:12px}}
.stat .value{{font-size:20px;font-weight:bold;margin-top:4px}}
.pos{{color:#e74c3c}} .neg{{color:#2ecc71}}
canvas{{margin-top:20px;background:#fff;border-radius:8px;padding:10px;box-shadow:0 1px 3px rgba(0,0,0,0.1)}}
</style>
</head><body>
<h2>双创等权指数</h2>
<p class="sub">创业板指 + 科创50 等权平均涨跌幅 | 归1复利净值 | {dates_raw[0]}~{dates_raw[-1]}</p>

<div class="stats">
  <div class="stat"><div class="label">最终净值</div><div class="value {nav_cls}">{final_nav:.4f}</div></div>
  <div class="stat"><div class="label">累计收益</div><div class="value {cum_cls}">{cum_ret:+.2f}%</div></div>
  <div class="stat"><div class="label">最大日涨幅</div><div class="value pos">{max_chg:+.2f}%</div></div>
  <div class="stat"><div class="label">最大日跌幅</div><div class="value neg">{min_chg:+.2f}%</div></div>
</div>

<canvas id="navChart" height="100"></canvas>
<canvas id="compChart" height="100"></canvas>
<script>
const dates = {json.dumps(dates)};
const navs = {json.dumps(navs)};
const avgChgs = {json.dumps(avg_chgs)};
const cybChgs = {json.dumps(cyb_chgs)};
const kcChgs = {json.dumps(kc_chgs)};

new Chart(document.getElementById('navChart'), {{
  type: 'line',
  data: {{labels: dates, datasets: [{{
    label: '归1净值', data: navs,
    borderColor: '#e74c3c', backgroundColor: 'rgba(231,76,60,0.08)',
    fill: true, tension: 0.3, pointRadius: 1.5, borderWidth: 2
  }}]}},
  options: {{
    plugins: {{title: {{display:true, text:'双创等权净值（归1）', font:{{size:14}}}}, legend:{{display:false}}}},
    scales: {{x:{{ticks:{{maxTicksLimit:12}}}}, y:{{title:{{display:true, text:'净值'}}}}}}
  }}
}});

// 创业板指 vs 科创50 每日涨跌幅对比
new Chart(document.getElementById('compChart'), {{
  type: 'line',
  data: {{labels: dates, datasets: [
    {{label: '创业板指%', data: cybChgs, borderColor: '#e74c3c', tension: 0.3, pointRadius: 1, borderWidth: 1.5}},
    {{label: '科创50%', data: kcChgs, borderColor: '#3498db', tension: 0.3, pointRadius: 1, borderWidth: 1.5}}
  ]}},
  options: {{
    plugins: {{title: {{display:true, text:'创业板指 vs 科创50 每日涨跌幅', font:{{size:14}}}}}},
    scales: {{x:{{ticks:{{maxTicksLimit:12}}}}, y:{{title:{{display:true, text:'%'}}}}}}
  }}
}});
</script>
</body></html>"""

# === 生成文件 ===
for pair in pairs:
    html = gen_spread_html(pair, rows1)
    path = os.path.join(BASE, pair['filename'])
    with open(path, 'w') as f:
        f.write(html)
    print(f"✅ {pair['filename']}")

sc_html = gen_sc_html(rows2)
sc_path = os.path.join(BASE, "shuangchuang.html")
with open(sc_path, 'w') as f:
    f.write(sc_html)
print(f"✅ shuangchuang.html")

print(f"\n全部生成在 {BASE}/")
