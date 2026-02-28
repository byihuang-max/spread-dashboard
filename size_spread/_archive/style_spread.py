#!/usr/bin/env python3
"""风格轧差 + 双创等权 + 经济敏感轧差 + 拥挤-反身性轧差
Sheet1: 风格轧差（中证红利-科创50, 微盘股-中证全指, 中证2000-沪深300）
Sheet2: 双创等权指数（创业板指+科创50 等权平均涨跌幅，归1净值）
Sheet3: 经济敏感轧差（申万有色+煤炭+钢铁 vs 食品饮料+医药，等权）
Sheet4: 拥挤-反身性轧差（申万31行业按20日成交额+波动率排序，Top6 vs Bottom6 等权）
输出: ~/Desktop/size_spread/style_spread.xlsx
"""
import statistics
import requests, datetime, time, os, sys

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

# === Tushare ===
TS_TOKEN = "8a2c71af4fbc6faf83da2ad4404c1c47f41983562cc9fb2fa6dd4fae"
TS_URL = "http://lianghua.nanyangqiankun.top"

def ts(api, params, fields=''):
    body = {"api_name": api, "token": TS_TOKEN, "params": params}
    if fields:
        body["fields"] = fields
    for attempt in range(3):
        try:
            r = requests.post(TS_URL, json=body, timeout=90)
            if not r.text.strip():
                print(f"  空响应，重试 {attempt+1}/3...")
                time.sleep(2); continue
            d = r.json()
            if d.get("data") and d["data"].get("fields") and d["data"].get("items"):
                return [dict(zip(d["data"]["fields"], row)) for row in d["data"]["items"]]
            return []
        except Exception as e:
            print(f"  请求失败: {e}，重试 {attempt+1}/3...")
            time.sleep(2)
    return []

# === 时间范围 ===
today = datetime.date.today()
start = (today - datetime.timedelta(days=400)).strftime("%Y%m%d")  # 近一年+余量
end = today.strftime("%Y%m%d")
start_long = (today - datetime.timedelta(days=400)).strftime("%Y%m%d")  # 拥挤度也拉一年

# ============================================================
# Part A: 拉指数数据（Sheet1 + Sheet2）
# ============================================================
all_codes = {
    '000922.CSI': '中证红利', '000688.SH': '科创50',
    '399303.SZ': '微盘股', '000985.CSI': '中证全指',
    '932000.CSI': '中证2000', '000300.SH': '沪深300',
    '399006.SZ': '创业板指',
}

index_data = {}
for code, name in all_codes.items():
    print(f"拉 {name}({code})...")
    d = ts('index_daily', {'ts_code': code, 'start_date': start, 'end_date': end},
           fields='trade_date,close,pct_chg')
    if d:
        print(f"  得到 {len(d)} 条")
        index_data[code] = {r['trade_date']: float(r['pct_chg']) for r in d}
    else:
        print(f"  ❌ 无数据")
        index_data[code] = {}

# ============================================================
# Part B: 拉申万31个一级行业（Sheet3 + Sheet4 共用）
# ============================================================
print("\n拉申万31个一级行业...")

sw_all_codes = {
    '801010.SI': '农林牧渔', '801030.SI': '基础化工', '801040.SI': '钢铁',
    '801050.SI': '有色金属', '801080.SI': '电子', '801880.SI': '汽车',
    '801110.SI': '家用电器', '801120.SI': '食品饮料', '801130.SI': '纺织服饰',
    '801140.SI': '轻工制造', '801150.SI': '医药生物', '801160.SI': '公用事业',
    '801170.SI': '交通运输', '801180.SI': '房地产', '801200.SI': '商贸零售',
    '801210.SI': '社会服务', '801780.SI': '银行', '801790.SI': '非银金融',
    '801230.SI': '综合', '801710.SI': '建筑材料', '801720.SI': '建筑装饰',
    '801730.SI': '电力设备', '801890.SI': '机械设备', '801740.SI': '国防军工',
    '801750.SI': '计算机', '801760.SI': '传媒', '801770.SI': '通信',
    '801950.SI': '煤炭', '801960.SI': '石油石化', '801970.SI': '环保',
    '801980.SI': '美容护理',
}

# code -> {date: {pct, amount}}
sw_daily_map = {}
for code, name in sw_all_codes.items():
    print(f"  {name}({code})...", end='', flush=True)
    d = ts('sw_daily', {'ts_code': code, 'start_date': start_long, 'end_date': end},
           fields='ts_code,trade_date,pct_change,amount')
    if d:
        print(f" {len(d)}条")
        sw_daily_map[code] = {}
        for r in d:
            dt = r['trade_date']
            pct = float(r['pct_change']) if r['pct_change'] is not None else 0.0
            amt = float(r['amount']) if r['amount'] is not None else 0.0
            sw_daily_map[code][dt] = {'pct': pct, 'amount': amt}
    else:
        print(" ❌")
        sw_daily_map[code] = {}

# ============================================================
# Sheet1: 风格轧差
# ============================================================
print("\n" + "=" * 50)
print("Sheet1: 风格轧差")
print("=" * 50)

pairs = [
    ('000922.CSI', '000688.SH', '中证红利', '科创50'),
    ('399303.SZ', '000985.CSI', '微盘股', '中证全指'),
    ('932000.CSI', '000300.SH', '中证2000', '沪深300'),
]

all_dates = set()
for c in all_codes:
    all_dates |= set(index_data.get(c, {}).keys())
all_dates = sorted(all_dates)

pair_navs = {}
for c1, c2, n1, n2 in pairs:
    label = f"{n1}-{n2}"
    d1, d2 = index_data.get(c1, {}), index_data.get(c2, {})
    common = sorted(set(d1.keys()) & set(d2.keys()))
    nav = 1.0
    result = {}
    for dt in common:
        spread = d1[dt] - d2[dt]
        nav *= (1 + spread / 100)
        result[dt] = {'chg1': d1[dt], 'chg2': d2[dt], 'spread': round(spread, 4), 'nav': round(nav, 6)}
    pair_navs[label] = result
    if common:
        final = result[common[-1]]
        print(f"  {label}: {len(common)}天, 净值{final['nav']:.4f}, 累计{(final['nav']-1)*100:+.2f}%")

# ============================================================
# Sheet2: 双创等权
# ============================================================
print("\n" + "=" * 50)
print("Sheet2: 双创等权")
print("=" * 50)

cyb = index_data.get('399006.SZ', {})
kc50 = index_data.get('000688.SH', {})
common_sc = sorted(set(cyb.keys()) & set(kc50.keys()))

sc_nav = 1.0
sc_data = {}
for dt in common_sc:
    avg_chg = (cyb[dt] + kc50[dt]) / 2
    sc_nav *= (1 + avg_chg / 100)
    sc_data[dt] = {'cyb_chg': cyb[dt], 'kc50_chg': kc50[dt], 'avg_chg': round(avg_chg, 4), 'nav': round(sc_nav, 6)}

if common_sc:
    final_sc = sc_data[common_sc[-1]]
    print(f"  双创等权: {len(common_sc)}天, 净值{final_sc['nav']:.4f}, 累计{(final_sc['nav']-1)*100:+.2f}%")

# ============================================================
# Sheet3: 经济敏感轧差
# 周期: 有色+煤炭+钢铁  vs  防御: 食品饮料+医药
# ============================================================
print("\n" + "=" * 50)
print("Sheet3: 经济敏感轧差")
print("=" * 50)

cycle_codes = ['801050.SI', '801950.SI', '801040.SI']   # 有色 煤炭 钢铁
defense_codes = ['801120.SI', '801150.SI']                # 食品饮料 医药

eco_dates_set = None
for code in cycle_codes + defense_codes:
    dates = set(sw_daily_map.get(code, {}).keys())
    eco_dates_set = dates if eco_dates_set is None else eco_dates_set & dates
eco_dates = sorted(eco_dates_set) if eco_dates_set else []
# 只取最近 ~45 个交易日（和 Sheet1 对齐）
eco_dates = [d for d in eco_dates if d >= start]

eco_nav = 1.0
eco_data = {}
for dt in eco_dates:
    cycle_pcts = [sw_daily_map[c][dt]['pct'] for c in cycle_codes if dt in sw_daily_map[c]]
    defense_pcts = [sw_daily_map[c][dt]['pct'] for c in defense_codes if dt in sw_daily_map[c]]
    if not cycle_pcts or not defense_pcts:
        continue
    cycle_chg = statistics.mean(cycle_pcts)
    defense_chg = statistics.mean(defense_pcts)
    spread = cycle_chg - defense_chg
    eco_nav *= (1 + spread / 100)
    eco_data[dt] = {
        'cycle_chg': round(cycle_chg, 4),
        'defense_chg': round(defense_chg, 4),
        'spread': round(spread, 4),
        'nav': round(eco_nav, 6),
    }
    for code in cycle_codes + defense_codes:
        eco_data[dt][code] = round(sw_daily_map[code][dt]['pct'], 4) if dt in sw_daily_map[code] else None

eco_dates = sorted(eco_data.keys())
if eco_dates:
    final_eco = eco_data[eco_dates[-1]]
    print(f"  周期-防御: {len(eco_dates)}天, 净值{final_eco['nav']:.4f}, 累计{(final_eco['nav']-1)*100:+.2f}%")

# ============================================================
# Sheet4: 动量轧差（原名：拥挤-反身性轧差）
#
# 【核心逻辑】
# 目标：衡量"近期热门行业"vs"近期冷门行业"的相对表现
#
# 1. 样本：申万31个一级行业指数（sw_daily）
#
# 2. 动量因子计算（每天滚动）：
#    对每个行业，取过去 LOOKBACK=20 个交易日（≈1个自然月）的数据：
#    - 指标A：20日平均成交额（amount均值）→ 按从高到低排名
#    - 指标B：20日波动率（pct_chg的标准差）→ 按从高到低排名
#    - 复合得分 = 成交额排名 + 波动率排名（越小=越"热"）
#
# 3. 分组：
#    - 高动量组 Top6：复合得分最小的6个行业（成交活跃+波动大）
#    - 低动量组 Bot6：复合得分最大的6个行业（成交冷清+波动小）
#    - 成分每天动态更新，不是固定的
#
# 4. 轧差计算：
#    每日轧差 = Top6等权平均涨跌幅 − Bot6等权平均涨跌幅
#    净值 = 归1复利累乘：∏(1 + 轧差/100)
#    正值 = 高动量行业跑赢低动量行业
#
# 5. "最近"的定义 = 20个交易日（LOOKBACK参数），约1个自然月
#    TOP_N = 6，即取排名前6和后6
# ============================================================
print("\n" + "=" * 50)
print("Sheet4: 动量轧差")
print("=" * 50)

# 找所有行业共同日期
crowd_dates_set = None
for code in sw_all_codes:
    if sw_daily_map.get(code):
        dates = set(sw_daily_map[code].keys())
        crowd_dates_set = dates if crowd_dates_set is None else crowd_dates_set & dates
crowd_all_dates = sorted(crowd_dates_set) if crowd_dates_set else []

LOOKBACK = 20
TOP_N = 6

crowd_data = {}
crowd_nav = 1.0

for i, dt in enumerate(crowd_all_dates):
    if i < LOOKBACK:
        continue

    window = crowd_all_dates[i - LOOKBACK:i]

    scores = {}
    for code in sw_all_codes:
        dm = sw_daily_map.get(code, {})
        amts = [dm[d]['amount'] for d in window if d in dm]
        pcts = [dm[d]['pct'] for d in window if d in dm]
        if len(amts) < 15 or len(pcts) < 15:
            continue
        scores[code] = {
            'avg_amt': statistics.mean(amts),
            'vol': statistics.stdev(pcts) if len(pcts) > 1 else 0,
        }

    if len(scores) < TOP_N * 2:
        continue

    codes_list = list(scores.keys())
    by_amt = sorted(codes_list, key=lambda c: scores[c]['avg_amt'], reverse=True)
    amt_rank = {c: r for r, c in enumerate(by_amt)}
    by_vol = sorted(codes_list, key=lambda c: scores[c]['vol'], reverse=True)
    vol_rank = {c: r for r, c in enumerate(by_vol)}
    composite = {c: amt_rank[c] + vol_rank[c] for c in codes_list}
    sorted_codes = sorted(codes_list, key=lambda c: composite[c])

    top_codes = sorted_codes[:TOP_N]
    bot_codes = sorted_codes[-TOP_N:]

    top_chgs = [sw_daily_map[c][dt]['pct'] for c in top_codes if dt in sw_daily_map[c]]
    bot_chgs = [sw_daily_map[c][dt]['pct'] for c in bot_codes if dt in sw_daily_map[c]]
    if not top_chgs or not bot_chgs:
        continue

    top_avg = statistics.mean(top_chgs)
    bot_avg = statistics.mean(bot_chgs)
    spread = top_avg - bot_avg
    crowd_nav *= (1 + spread / 100)

    crowd_data[dt] = {
        'top_chg': round(top_avg, 4),
        'bot_chg': round(bot_avg, 4),
        'spread': round(spread, 4),
        'nav': round(crowd_nav, 6),
        'top_names': ','.join([sw_all_codes[c] for c in top_codes]),
        'bot_names': ','.join([sw_all_codes[c] for c in bot_codes]),
    }

crowd_dates = sorted(crowd_data.keys())
if crowd_dates:
    final_cr = crowd_data[crowd_dates[-1]]
    print(f"  高拥挤-低拥挤: {len(crowd_dates)}天, 净值{final_cr['nav']:.4f}, 累计{(final_cr['nav']-1)*100:+.2f}%")
    print(f"  最新Top6(高拥挤): {final_cr['top_names']}")
    print(f"  最新Bot6(低拥挤): {final_cr['bot_names']}")

# ============================================================
# 写 Excel
# ============================================================
print("\n写入 Excel...")
xlsx_path = os.path.expanduser("~/Desktop/size_spread/style_spread.xlsx")
wb = openpyxl.Workbook()

# --- Sheet1 ---
ws1 = wb.active
ws1.title = "风格轧差"
headers = ['日期']
for c1, c2, n1, n2 in pairs:
    headers += [f'{n1}涨跌幅%', f'{n2}涨跌幅%', f'{n1}-{n2}轧差%', f'{n1}-{n2}净值']
ws1.append(headers)
for dt in all_dates:
    row = [dt]
    for c1, c2, n1, n2 in pairs:
        label = f"{n1}-{n2}"
        r = pair_navs.get(label, {}).get(dt)
        if r:
            row += [round(r['chg1'], 4), round(r['chg2'], 4), r['spread'], r['nav']]
        else:
            row += [None, None, None, None]
    ws1.append(row)

# --- Sheet2 ---
ws2 = wb.create_sheet("双创等权")
ws2.append(['日期', '创业板指涨跌幅%', '科创50涨跌幅%', '等权平均涨跌幅%', '归1净值'])
for dt in common_sc:
    r = sc_data[dt]
    ws2.append([dt, round(r['cyb_chg'], 4), round(r['kc50_chg'], 4), r['avg_chg'], r['nav']])

# --- Sheet3 ---
ws3 = wb.create_sheet("经济敏感轧差")
cycle_names = [sw_all_codes[c] for c in cycle_codes]
defense_names = [sw_all_codes[c] for c in defense_codes]
eco_headers = ['日期'] + [f'{n}%' for n in cycle_names] + ['周期等权%'] + \
              [f'{n}%' for n in defense_names] + ['防御等权%', '周期-防御轧差%', '轧差净值']
ws3.append(eco_headers)
for dt in eco_dates:
    r = eco_data[dt]
    row = [dt]
    for code in cycle_codes:
        row.append(r.get(code))
    row.append(r['cycle_chg'])
    for code in defense_codes:
        row.append(r.get(code))
    row += [r['defense_chg'], r['spread'], r['nav']]
    ws3.append(row)

# --- Sheet4 ---
ws4 = wb.create_sheet("拥挤-反身性轧差")
ws4.append(['日期', '高拥挤Top6等权%', '低拥挤Bot6等权%', '高-低轧差%', '轧差净值',
            'Top6行业', 'Bottom6行业'])
for dt in crowd_dates:
    r = crowd_data[dt]
    ws4.append([dt, r['top_chg'], r['bot_chg'], r['spread'], r['nav'],
                r['top_names'], r['bot_names']])

wb.save(xlsx_path)
print(f"\n✅ 已保存: {xlsx_path}")
print(f"   Sheet1: 风格轧差 ({len(all_dates)} 行 × 3对)")
print(f"   Sheet2: 双创等权 ({len(common_sc)} 行)")
print(f"   Sheet3: 经济敏感轧差 ({len(eco_dates)} 行)")
print(f"   Sheet4: 拥挤-反身性轧差 ({len(crowd_dates)} 行)")

# ============================================================
# 输出 JSON（供 dashboard 使用）
# ============================================================
import json

dashboard_data_dir = os.path.expanduser("~/Desktop/gamt-dashboard/data")
os.makedirs(dashboard_data_dir, exist_ok=True)

# --- 风格轧差 JSON（Sheet1 + Sheet2 + Sheet3 + Sheet4 合并） ---
style_json = {
    "update_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    # Sheet1: 风格轧差
    "style_spread": {
        "pairs": [{"long": n1, "short": n2} for _, _, n1, n2 in pairs],
        "dates": [],
        "data": {}  # label -> {navs:[], spreads:[]}
    },
    # Sheet2: 双创等权
    "dual_innovation": {
        "dates": [],
        "navs": [],
        "cyb_chg": [],
        "kc50_chg": [],
    },
    # Sheet3: 经济敏感轧差
    "eco_sensitive": {
        "cycle_names": [sw_all_codes[c] for c in cycle_codes],
        "defense_names": [sw_all_codes[c] for c in defense_codes],
        "dates": [],
        "cycle_chg": [],
        "defense_chg": [],
        "spreads": [],
        "navs": [],
    },
    # Sheet4: 拥挤-反身性轧差
    "crowding": {
        "dates": [],
        "top_chg": [],
        "bot_chg": [],
        "spreads": [],
        "navs": [],
        "top_names": [],
        "bot_names": [],
    },
}

# Sheet1 数据
for c1, c2, n1, n2 in pairs:
    label = f"{n1}-{n2}"
    pdata = pair_navs.get(label, {})
    common = sorted(pdata.keys())
    style_json["style_spread"]["data"][label] = {
        "dates": [d[4:6]+'/'+d[6:8] for d in common],
        "navs": [pdata[d]['nav'] for d in common],
        "spreads": [pdata[d]['spread'] for d in common],
    }

# Sheet2 数据
for dt in common_sc:
    r = sc_data[dt]
    style_json["dual_innovation"]["dates"].append(dt[4:6]+'/'+dt[6:8])
    style_json["dual_innovation"]["navs"].append(r['nav'])
    style_json["dual_innovation"]["cyb_chg"].append(r['cyb_chg'])
    style_json["dual_innovation"]["kc50_chg"].append(r['kc50_chg'])

# Sheet3 数据
for dt in eco_dates:
    r = eco_data[dt]
    style_json["eco_sensitive"]["dates"].append(dt[4:6]+'/'+dt[6:8])
    style_json["eco_sensitive"]["cycle_chg"].append(r['cycle_chg'])
    style_json["eco_sensitive"]["defense_chg"].append(r['defense_chg'])
    style_json["eco_sensitive"]["spreads"].append(r['spread'])
    style_json["eco_sensitive"]["navs"].append(r['nav'])

# Sheet4 数据
for dt in crowd_dates:
    r = crowd_data[dt]
    style_json["crowding"]["dates"].append(dt[4:6]+'/'+dt[6:8])
    style_json["crowding"]["top_chg"].append(r['top_chg'])
    style_json["crowding"]["bot_chg"].append(r['bot_chg'])
    style_json["crowding"]["spreads"].append(r['spread'])
    style_json["crowding"]["navs"].append(r['nav'])
    style_json["crowding"]["top_names"].append(r['top_names'])
    style_json["crowding"]["bot_names"].append(r['bot_names'])

json_path = os.path.join(dashboard_data_dir, "style_spread.json")
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(style_json, f, ensure_ascii=False, indent=2)
print(f"   JSON: {json_path}")

# ============================================================
# 生成静态 HTML（更新 size_spread.html）
# ============================================================
import json as _json

def fmt_date(d):
    return d[4:6] + '/' + d[6:8]

# 准备各组数据
# --- 风格轧差（3对）---
style_series = {}
for c1, c2, n1, n2 in pairs:
    label = f"{n1}-{n2}"
    pdata = pair_navs.get(label, {})
    common = sorted(pdata.keys())
    style_series[label] = {
        'dates': [fmt_date(d) for d in common],
        'navs': [pdata[d]['nav'] for d in common],
        'spreads': [pdata[d]['spread'] for d in common],
    }

# --- 双创等权 ---
di_dates = [fmt_date(d) for d in common_sc]
di_navs = [sc_data[d]['nav'] for d in common_sc]

# --- 经济敏感 ---
eco_d = [fmt_date(d) for d in eco_dates]
eco_n = [eco_data[d]['nav'] for d in eco_dates]
eco_s = [eco_data[d]['spread'] for d in eco_dates]
eco_final_nav = eco_n[-1] if eco_n else 1.0

# --- 拥挤度 ---
cr_d = [fmt_date(d) for d in crowd_dates]
cr_n = [crowd_data[d]['nav'] for d in crowd_dates]
cr_s = [crowd_data[d]['spread'] for d in crowd_dates]
cr_final_nav = cr_n[-1] if cr_n else 1.0
cr_top = crowd_data[crowd_dates[-1]]['top_names'].replace(',', ' · ') if crowd_dates else '-'
cr_bot = crowd_data[crowd_dates[-1]]['bot_names'].replace(',', ' · ') if crowd_dates else '-'

# 第一对风格轧差的日期（用于 x 轴）
first_label = list(style_series.keys())[0]

html = f'''<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>风格轧差看板</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
body{{font-family:'PingFang SC',sans-serif;max-width:1100px;margin:30px auto;padding:0 20px;background:#fafafa}}
h2{{text-align:center;color:#333;margin-bottom:5px}}
p.sub{{text-align:center;color:#888;font-size:13px;margin-top:0}}
.cards{{display:flex;justify-content:center;gap:20px;margin:15px 0;flex-wrap:wrap}}
.card{{background:#fff;padding:12px 18px;border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,0.1);min-width:140px;text-align:center}}
.card .label{{color:#888;font-size:11px}}
.card .value{{font-size:18px;font-weight:bold;margin-top:3px}}
.card .detail{{font-size:11px;color:#999;margin-top:2px}}
.pos{{color:#e74c3c}} .neg{{color:#2ecc71}}
.section{{margin-top:35px}}
.section h3{{color:#555;font-size:15px;border-bottom:1px solid #eee;padding-bottom:5px}}
.row{{display:flex;gap:15px;margin-top:10px}}
.row canvas{{flex:1;background:#fff;border-radius:8px;padding:8px;box-shadow:0 1px 3px rgba(0,0,0,0.1)}}
.tag-row{{display:flex;gap:8px;flex-wrap:wrap;margin:8px 0}}
.tag{{background:#fff3e0;color:#e65100;padding:3px 10px;border-radius:12px;font-size:12px}}
.tag.cool{{background:#e3f2fd;color:#1565c0}}
</style>
</head><body>

<h2>风格轧差看板</h2>
<p class="sub">数据来源: Tushare 申万行业指数 | 最后更新: {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}</p>

<!-- ===== 经济敏感轧差 ===== -->
<div class="section">
<h3>📊 经济敏感轧差（有色+煤炭+钢铁 vs 食品饮料+医药）</h3>
<div class="cards">
  <div class="card"><div class="label">轧差净值</div><div class="value {("pos" if eco_final_nav>=1 else "neg")}">{eco_final_nav:.4f}</div></div>
  <div class="card"><div class="label">累计收益</div><div class="value {("pos" if eco_final_nav>=1 else "neg")}">{(eco_final_nav-1)*100:+.2f}%</div></div>
  <div class="card"><div class="label">观察天数</div><div class="value">{len(eco_dates)}</div></div>
</div>
<div class="row">
  <canvas id="ecoNavChart" height="120"></canvas>
  <canvas id="ecoSpreadChart" height="120"></canvas>
</div>
</div>

<!-- ===== 拥挤-反身性轧差 ===== -->
<div class="section">
<h3>🔥 拥挤-反身性轧差（高拥挤Top6 vs 低拥挤Bot6，20日滚动）</h3>
<div class="cards">
  <div class="card"><div class="label">轧差净值</div><div class="value {("pos" if cr_final_nav>=1 else "neg")}">{cr_final_nav:.4f}</div></div>
  <div class="card"><div class="label">累计收益</div><div class="value {("pos" if cr_final_nav>=1 else "neg")}">{(cr_final_nav-1)*100:+.2f}%</div></div>
  <div class="card"><div class="label">观察天数</div><div class="value">{len(crowd_dates)}</div></div>
</div>
<p style="font-size:12px;color:#666;margin:5px 0 0 0">最新高拥挤 Top6：</p>
<div class="tag-row">''' + ''.join(f'<span class="tag">{n}</span>' for n in cr_top.split(' · ')) + '''</div>
<p style="font-size:12px;color:#666;margin:5px 0 0 0">最新低拥挤 Bot6：</p>
<div class="tag-row">''' + ''.join(f'<span class="tag cool">{n}</span>' for n in cr_bot.split(' · ')) + f'''</div>
<div class="row">
  <canvas id="crowdNavChart" height="120"></canvas>
  <canvas id="crowdSpreadChart" height="120"></canvas>
</div>
</div>

<!-- ===== 风格轧差净值 ===== -->
<div class="section">
<h3>📈 风格轧差净值（归1）</h3>
<div class="row">
  <canvas id="styleNavChart" height="140"></canvas>
</div>
</div>

<!-- ===== 双创等权 ===== -->
<div class="section">
<h3>🚀 双创等权净值（创业板指+科创50 等权）</h3>
<div class="row">
  <canvas id="dualChart" height="120"></canvas>
</div>
</div>

<script>
const eco_dates = {_json.dumps(eco_d)};
const eco_navs = {_json.dumps(eco_n)};
const eco_spreads = {_json.dumps(eco_s)};
const cr_dates = {_json.dumps(cr_d)};
const cr_navs = {_json.dumps(cr_n)};
const cr_spreads = {_json.dumps(cr_s)};
const style_data = {_json.dumps(style_series)};
const di_dates = {_json.dumps(di_dates)};
const di_navs = {_json.dumps(di_navs)};

const lineOpts = (title) => ({{
  plugins:{{title:{{display:true,text:title,font:{{size:13}}}},legend:{{display:false}}}},
  scales:{{x:{{ticks:{{maxTicksLimit:10}}}},y:{{title:{{display:true,text:'净值'}}}}}}
}});
const barOpts = (title) => ({{
  plugins:{{title:{{display:true,text:title,font:{{size:13}}}},legend:{{display:false}}}},
  scales:{{x:{{ticks:{{maxTicksLimit:10}}}},y:{{title:{{display:true,text:'%'}}}}}}
}});

// 经济敏感
new Chart(document.getElementById('ecoNavChart'),{{type:'line',data:{{labels:eco_dates,datasets:[{{
  data:eco_navs,borderColor:'#e67e22',backgroundColor:'rgba(230,126,34,0.08)',fill:true,tension:0.3,pointRadius:1,borderWidth:2
}}]}},options:lineOpts('周期-防御 净值')}});

new Chart(document.getElementById('ecoSpreadChart'),{{type:'bar',data:{{labels:eco_dates,datasets:[{{
  data:eco_spreads,backgroundColor:eco_spreads.map(v=>v>=0?'rgba(231,76,60,0.6)':'rgba(52,152,219,0.6)'),borderRadius:2
}}]}},options:barOpts('周期-防御 每日轧差%')}});

// 拥挤度
new Chart(document.getElementById('crowdNavChart'),{{type:'line',data:{{labels:cr_dates,datasets:[{{
  data:cr_navs,borderColor:'#c0392b',backgroundColor:'rgba(192,57,43,0.08)',fill:true,tension:0.3,pointRadius:1,borderWidth:2
}}]}},options:lineOpts('高拥挤-低拥挤 净值')}});

new Chart(document.getElementById('crowdSpreadChart'),{{type:'bar',data:{{labels:cr_dates,datasets:[{{
  data:cr_spreads,backgroundColor:cr_spreads.map(v=>v>=0?'rgba(192,57,43,0.6)':'rgba(41,128,185,0.6)'),borderRadius:2
}}]}},options:barOpts('高拥挤-低拥挤 每日轧差%')}});

// 风格轧差（多线）
const colors = ['#e74c3c','#3498db','#2ecc71'];
const styleDs = [];
let ci = 0;
for (const label in style_data) {{
  styleDs.push({{label:label,data:style_data[label].navs,borderColor:colors[ci%3],backgroundColor:'transparent',tension:0.3,pointRadius:1,borderWidth:2}});
  ci++;
}}
const firstKey = Object.keys(style_data)[0];
new Chart(document.getElementById('styleNavChart'),{{type:'line',data:{{labels:style_data[firstKey].dates,datasets:styleDs}},options:{{
  plugins:{{title:{{display:false}},legend:{{display:true,position:'top'}}}},
  scales:{{x:{{ticks:{{maxTicksLimit:10}}}},y:{{title:{{display:true,text:'净值'}}}}}}
}}}});

// 双创等权
new Chart(document.getElementById('dualChart'),{{type:'line',data:{{labels:di_dates,datasets:[{{
  data:di_navs,borderColor:'#9b59b6',backgroundColor:'rgba(155,89,182,0.08)',fill:true,tension:0.3,pointRadius:1,borderWidth:2
}}]}},options:lineOpts('双创等权净值')}});
</script>
</body></html>'''

html_path = os.path.expanduser("~/Desktop/size_spread/风格轧差看板.html")
with open(html_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"   HTML: {html_path}")
